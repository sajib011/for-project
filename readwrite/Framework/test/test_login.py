import unittest

from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
import time
import pytest
import openpyxl
# from Framework.pages.login_page import LoginPage
from selenium.webdriver.common.by import By
# from Framework.utils import excel_utils

# Read data and implementing in test


class LoginTest(unittest.TestCase):

    def test_in(self):
        driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))

        driver.implicitly_wait(3)
        driver.maximize_window()

        time.sleep(3)
        # read from excel file data to website
        # write data to excel file

        wb=openpyxl.load_workbook("D:\\software testing class\\demo part\\Framework\\data\\login_data.xlsx")
        sh1=wb['Sheet1']
        # row=sh1.max_row
        # row1=row+1
        # print(row1)
        # column=sh1.max_column
        for i in range(2, 6):
            driver.get("https://admin-demo.nopcommerce.com/login?ReturnUrl=%2Fadmin%2F")
            username_data = sh1.cell(i,1).value
            username_field = driver.find_element(By.ID, "Email")
            username_field.clear()
            username_field.send_keys(username_data)
            time.sleep(3)
            password_data = sh1.cell(i,2).value
            password_field = driver.find_element(By.ID, "Password")
            password_field.clear()
            password_field.send_keys(password_data)
            time.sleep(3)
            login_button = driver.find_element(By.XPATH,"/html/body/div[6]/div/div/div/div/div[2]/div[1]/div/form/div[3]/button")
            login_button.click()

            if driver.current_url=="https://admin-demo.nopcommerce.com/admin/":
                sh1.cell(i,4, value='Login')
            else:
                sh1.cell(i, 4, value='Not Login')
        wb.save("newdata.xlsx")



        driver.close()
